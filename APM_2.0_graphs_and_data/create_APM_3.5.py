"""
Generate APM_3.5_Summary.xlsx from Analysis_Scripts/data/summary_per_group_dev.csv
Columns: APM Prime Bit Number, Matrix Size (n), Hits, Total, Minors Tested, Total Zero Minors, Hit Ratio
Grouped by prime-bit group, each group gets a title + header + data rows (one row per deviation).
"""

import csv
import math
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── paths ────────────────────────────────────────────────────────────────────
BASE   = os.path.dirname(os.path.abspath(__file__))
ROOT   = os.path.dirname(BASE)
CSV_IN = os.path.join(ROOT, "Analysis_Scripts", "data", "summary_per_group_dev.csv")
XLSX   = os.path.join(BASE, "APM_3.5_Summary.xlsx")

# ── colour palette ────────────────────────────────────────────────────────────
CLR_TITLE_BG   = "1F3864"   # dark navy   – title rows
CLR_TITLE_FG   = "FFFFFF"   # white
CLR_HEADER_BG  = "2E75B6"   # mid-blue    – column headers
CLR_HEADER_FG  = "FFFFFF"
CLR_ROW_ODD    = "D9E2F3"   # light blue  – alternating data rows
CLR_ROW_EVEN   = "FFFFFF"
CLR_HIT100     = "E2EFDA"   # pale green  – 100 % hit ratio row
CLR_BORDER     = "9DC3E6"   # border colour

# ── border helper ─────────────────────────────────────────────────────────────
thin_side  = Side(style="thin",   color=CLR_BORDER)
thick_side = Side(style="medium", color="1F3864")

def thin_border():
    return Border(left=thin_side, right=thin_side,
                  top=thin_side,  bottom=thin_side)

def bottom_thick():
    return Border(left=thin_side,  right=thin_side,
                  top=thin_side,   bottom=thick_side)

# ── read & group CSV ──────────────────────────────────────────────────────────
rows_by_group: dict[int, list[dict]] = {}
with open(CSV_IN, newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        g = int(row["group"])
        rows_by_group.setdefault(g, []).append(row)

# ── workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "APM Summary 3.5"

# column widths  A … G
col_widths = [22, 16, 12, 10, 26, 24, 14]
COLS = 7

for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# freeze the top row once we place headers (not used – no persistent header)

current_row = 1

for group in sorted(rows_by_group.keys()):
    group_rows = rows_by_group[group]

    # grab prime from first entry in group
    prime_val = int(group_rows[0]["prime"])
    prime_fmt = f"{prime_val:,}"
    # matrix size n  (from 'matrices' column – that is the count tested, not n)
    # The APM 2.0 style uses "Matrix size n = <value>" in the title.
    # In the CSV, 'matrices' is the number of matrices tested (always 100).
    # We derive n from the group (prime bit) following the pattern in APM 2.0:
    #   n ≈ group - 10  (e.g. group 25 → n=15, group 26 → n=16, …)
    # Use that formula; override from 'matrices' is unreliable for this.
    n_val = group - 10

    # ── title row ────────────────────────────────────────────────────────────
    title_text = (
        f"APM Data — Prime Bit {group}    "
        f"P = {prime_fmt}    "
        f"Matrix size n = {n_val}"
    )
    ws.merge_cells(
        start_row=current_row, start_column=1,
        end_row=current_row,   end_column=COLS
    )
    title_cell = ws.cell(row=current_row, column=1, value=title_text)
    title_cell.font      = Font(bold=True, color=CLR_TITLE_FG, size=12, name="Calibri")
    title_cell.fill      = PatternFill("solid", fgColor=CLR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                      indent=1)
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    # ── column header row ────────────────────────────────────────────────────
    headers = [
        "APM Prime Bit",
        "Matrix Size (n×n)",
        "Hits",
        "Total",
        "Minors Tested",
        "Total Zero Minors",
        "Hit Ratio",
    ]
    for col_idx, hdr in enumerate(headers, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=hdr)
        c.font      = Font(bold=True, color=CLR_HEADER_FG, size=10, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── data rows ────────────────────────────────────────────────────────────
    for i, row in enumerate(group_rows):
        dev          = int(row["dev"])
        total_hits   = int(row["total_hits"])
        matrices     = int(row["matrices"])        # always 100
        minors_mean  = float(row["minors_mean"])
        zero_minors  = int(row["zero_minors"])
        hit_ratio    = float(row["hit_ratio"])

        total_minors = int(round(minors_mean * matrices))
        matrix_size_label = f"{dev+2}×{dev+2}"   # minor size (as in APM 2.0)
        hit_ratio_pct = f"{hit_ratio*100:.2f}%"
        hits_label    = f"{total_hits}/{matrices}"

        is_hit100 = (hit_ratio >= 1.0)
        is_odd    = (i % 2 == 0)

        if is_hit100:
            row_bg = CLR_HIT100
        elif is_odd:
            row_bg = CLR_ROW_ODD
        else:
            row_bg = CLR_ROW_EVEN

        row_fill = PatternFill("solid", fgColor=row_bg)
        row_font = Font(name="Calibri", size=10)

        values = [
            group,
            matrix_size_label,
            hits_label,
            matrices,
            total_minors,
            zero_minors,
            hit_ratio_pct,
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.fill      = row_fill
            c.font      = row_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            # last data row of group gets a thicker bottom border
            if i == len(group_rows) - 1:
                c.border = bottom_thick()
            else:
                c.border = thin_border()

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # blank spacer between groups
    current_row += 1

# ── freeze top row (sheet header) ────────────────────────────────────────────
# (no fixed top-of-sheet header in this layout; nothing to freeze)

# ── auto-filter on first data header (row 2) — skip, mixed groups ─────────────
# kept clean: no auto-filter to avoid confusion across groups

wb.save(XLSX)
print(f"Saved → {XLSX}")
